"""
Run Optimization Callbacks
Handles table display, editing, validation, and Bayesian optimization
"""

from dash import callback, Input, Output, State, html, no_update, ctx, dcc
from dash.exceptions import PreventUpdate
import dash_bootstrap_components as dbc
from dash import dash_table
import pandas as pd
import numpy as np
import os

from domain_storage import DomainStorage
from utils.BoFire import (
    create_bofire_domain_from_store, 
    bayesian_optimization,
    create_acquisition_function_from_name,
    get_optimization_type
)
from utils.safe_excel import safe_excel_save, safe_excel_read
from config_path import EXCEL_FOLDER
from callbacks.opti_param_callbacks.constraints_callbacks import validate_and_adjust_suggestion


# ===== LOAD AND DISPLAY TABLE =====

@callback(
    [Output('experiment-table-container', 'children'),
     Output('run-status-alert', 'children'),
     Output('run-status-alert', 'is_open'),
     Output('run-status-alert', 'color')],
    [Input('current-excel-file', 'data'),
     Input('url', 'pathname')],
    prevent_initial_call=False
)
def load_experiment_table(excel_file, pathname):
    """Load Excel file and display as editable DataTable"""
    
    if pathname != '/Opt-run':
        raise PreventUpdate
    
    if not excel_file:
        return (
            html.Div([
                html.I(className="bi bi-exclamation-circle", style={"fontSize": "2rem", "color": "#6c757d"}),
                html.P("No project selected", className="mt-2 text-muted"),
                html.A("← Go to Home", href="/Opt-home", className="btn btn-outline-primary btn-sm")
            ], className="text-center py-5"),
            "Please create or select a project first",
            True,
            "warning"
        )
    
    try:
        file_path = os.path.join(EXCEL_FOLDER, excel_file)
        
        if not os.path.exists(file_path):
            return (
                html.Div(f"File not found: {excel_file}"),
                f"File not found: {excel_file}",
                True,
                "danger"
            )
        
        # === SAFE READ with automatic backup recovery ===
        df, read_msg = safe_excel_read(file_path)
        if df is None:
            return html.Div(f"Error: {read_msg}"), read_msg, True, "danger"
        if "backup" in read_msg:
            print(f"⚠️ {read_msg}")
        
        domain_data = DomainStorage.load_domain(excel_file)
        if not domain_data:
            return (
                html.Div("Domain configuration missing"),
                "Domain not configured",
                True,
                "warning"
            )
        
        param_names = domain_data.get('metadata', {}).get('parameter_names', [])
        obj_names = domain_data.get('metadata', {}).get('objective_names', [])
        
        columns = [{'name': col, 'id': col, 'editable': True} for col in df.columns]
        
        style_data_conditional = []
        for param in param_names:
            if param in df.columns:
                style_data_conditional.append({
                    'if': {'column_id': param},
                    'backgroundColor': 'rgba(0, 123, 255, 0.1)'
                })
        
        for obj in obj_names:
            if obj in df.columns:
                style_data_conditional.append({
                    'if': {'column_id': obj},
                    'backgroundColor': 'rgba(40, 167, 69, 0.1)'
                })
        
        if 'Point type' in df.columns:
            style_data_conditional.append({
                'if': {'column_id': 'Point type'},
                'backgroundColor': 'rgba(255, 107, 53, 0.1)'
            })
        
        for obj in obj_names:
            style_data_conditional.append({
                'if': {
                    'filter_query': f'{{{obj}}} is blank',
                    'column_id': obj
                },
                'backgroundColor': 'rgba(255, 193, 7, 0.3)',
                'border': '2px solid #ffc107'
            })
        
        table = dash_table.DataTable(
            id='experiment-datatable',
            columns=columns,
            data=df.to_dict('records'),
            editable=True,
            row_deletable=True,
            style_table={
                'overflowX': 'auto',
                'borderRadius': '8px',
                'border': '1px solid #e0e0e0'
            },
            style_header={
                'backgroundColor': '#f8f9fa',
                'fontWeight': 'bold',
                'textAlign': 'center',
                'padding': '12px',
                'borderBottom': '2px solid #dee2e6'
            },
            style_cell={
                'textAlign': 'center',
                'padding': '10px',
                'fontSize': '0.875rem',
                'minWidth': '100px'
            },
            style_data_conditional=style_data_conditional,
            page_size=20,
            export_format='xlsx'
        )
        
        incomplete = 0
        for _, row in df.iterrows():
            for obj in obj_names:
                if obj in row and (pd.isna(row[obj]) or row[obj] == ''):
                    incomplete += 1
                    break
        
        if incomplete > 0:
            status = f"⚠️ {incomplete} experiment(s) need results. Fill in objective values to enable optimization."
            color = "warning"
        else:
            status = "✅ All experiments have results. Ready for Bayesian optimization!"
            color = "success"
        
        # Add backup recovery warning if applicable
        if "backup" in read_msg:
            status = f"⚠️ {read_msg}. " + status
            color = "warning"
        
        return table, status, True, color
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return html.Div(f"Error: {str(e)}"), f"Failed to load: {str(e)}", True, "danger"


# ===== SAVE TABLE CHANGES =====

@callback(
    [Output('save-status', 'children'),
     Output('save-status', 'is_open')],
    Input('save-table-btn', 'n_clicks'),
    [State('experiment-datatable', 'data'),
     State('current-excel-file', 'data')],
    prevent_initial_call=True
)
def save_table_changes(n_clicks, table_data, excel_file):
    """Save edited table back to Excel"""
    
    if not n_clicks or not table_data or not excel_file:
        raise PreventUpdate
    
    file_path = os.path.join(EXCEL_FOLDER, excel_file)
    df = pd.DataFrame(table_data)
    
    def write_formatted(path):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Experiments')
            
            domain_data = DomainStorage.load_domain(excel_file)
            if domain_data:
                from openpyxl.styles import Font, PatternFill, Alignment
                worksheet = writer.sheets['Experiments']
                
                param_names = domain_data.get('metadata', {}).get('parameter_names', [])
                obj_names = domain_data.get('metadata', {}).get('objective_names', [])
                
                for i, cell in enumerate(worksheet[1]):
                    col_name = df.columns[i]
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center')
                    
                    if col_name in param_names:
                        cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
                    elif col_name in obj_names:
                        cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                    elif col_name == 'Point type':
                        cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
    
    # === SAFE SAVE with atomic write + backup ===
    success, msg = safe_excel_save(file_path, write_formatted)
    
    if success:
        return dbc.Alert("✅ Table saved successfully!", color="success"), True
    else:
        return dbc.Alert(f"❌ {msg}", color="danger"), True


# ===== ADD NEW ROW =====

@callback(
    Output('experiment-datatable', 'data', allow_duplicate=True),
    Input('add-row-btn', 'n_clicks'),
    State('experiment-datatable', 'data'),
    State('experiment-datatable', 'columns'),
    prevent_initial_call=True
)
def add_new_row(n_clicks, data, columns):
    """Add a new empty row to the table"""
    
    if not n_clicks:
        raise PreventUpdate
    
    new_row = {}
    for col in columns:
        if col['id'] == 'Point type':
            new_row[col['id']] = 'BO'
        else:
            new_row[col['id']] = ''
    
    data.append(new_row)
    return data


# ===== ADD/DELETE COLUMN MODALS =====

@callback(
    [Output('opti-add-column-modal', 'is_open'),
     Output('opti-delete-column-modal', 'is_open')],
    [Input('add-column-btn', 'n_clicks'),
     Input('opti-cancel-add-column', 'n_clicks'),
     Input('opti-confirm-add-column', 'n_clicks'),
     Input('delete-column-btn', 'n_clicks'),
     Input('opti-cancel-delete-column', 'n_clicks'),
     Input('opti-confirm-delete-column', 'n_clicks')],
    [State('opti-add-column-modal', 'is_open'),
     State('opti-delete-column-modal', 'is_open')],
    prevent_initial_call=True
)
def toggle_modals(add_open, add_cancel, add_confirm, del_open, del_cancel, del_confirm, add_is_open, del_is_open):
    """Toggle add/delete column modals"""
    triggered = ctx.triggered_id
    
    if triggered == 'add-column-btn':
        return True, False
    elif triggered in ['opti-cancel-add-column', 'opti-confirm-add-column']:
        return False, False
    elif triggered == 'delete-column-btn':
        return False, True
    elif triggered in ['opti-cancel-delete-column', 'opti-confirm-delete-column']:
        return False, False
    
    return add_is_open, del_is_open


@callback(
    [Output('experiment-datatable', 'columns', allow_duplicate=True),
     Output('experiment-datatable', 'data', allow_duplicate=True),
     Output('opti-new-column-name', 'value'),
     Output('opti-new-column-default', 'value')],
    Input('opti-confirm-add-column', 'n_clicks'),
    [State('opti-new-column-name', 'value'),
     State('opti-new-column-default', 'value'),
     State('experiment-datatable', 'columns'),
     State('experiment-datatable', 'data')],
    prevent_initial_call=True
)
def add_new_column(n_clicks, col_name, default_value, columns, data):
    """Add a new column to the table"""
    
    if not n_clicks or not col_name or not col_name.strip():
        raise PreventUpdate
    
    col_name = col_name.strip()
    
    existing_cols = [c['id'] for c in columns]
    if col_name in existing_cols:
        raise PreventUpdate
    
    new_col = {
        'name': col_name,
        'id': col_name,
        'editable': True,
        'type': 'text'
    }
    columns.append(new_col)
    
    default = default_value if default_value else ''
    for row in data:
        row[col_name] = default
    
    return columns, data, '', ''


@callback(
    Output('opti-column-to-delete', 'options'),
    Input('opti-delete-column-modal', 'is_open'),
    State('experiment-datatable', 'columns'),
    prevent_initial_call=True
)
def populate_delete_column_dropdown(is_open, columns):
    """Populate dropdown with column names"""
    if not is_open or not columns:
        raise PreventUpdate
    
    return [{'label': c['name'], 'value': c['id']} for c in columns]


@callback(
    [Output('experiment-datatable', 'columns', allow_duplicate=True),
     Output('experiment-datatable', 'data', allow_duplicate=True)],
    Input('opti-confirm-delete-column', 'n_clicks'),
    [State('opti-column-to-delete', 'value'),
     State('experiment-datatable', 'columns'),
     State('experiment-datatable', 'data')],
    prevent_initial_call=True
)
def delete_column(n_clicks, col_to_delete, columns, data):
    """Delete a column from the table"""
    
    if not n_clicks or not col_to_delete:
        raise PreventUpdate
    
    columns = [c for c in columns if c['id'] != col_to_delete]
    
    for row in data:
        if col_to_delete in row:
            del row[col_to_delete]
    
    return columns, data


# ===== VALIDATE AND ENABLE OPTIMIZATION =====

@callback(
    Output('run-bo-btn', 'disabled'),
    Input('experiment-datatable', 'data'),
    State('current-excel-file', 'data'),
    prevent_initial_call=True
)
def validate_for_optimization(table_data, excel_file):
    """Check if all objective values are filled to enable BO button"""
    
    if not table_data or not excel_file:
        return True
    
    try:
        domain_data = DomainStorage.load_domain(excel_file)
        if not domain_data:
            return True
        
        obj_names = domain_data.get('metadata', {}).get('objective_names', [])
        
        if not obj_names:
            return True
        
        for row in table_data:
            for obj in obj_names:
                if obj in row:
                    val = row[obj]
                    if val is None or val == '' or (isinstance(val, float) and np.isnan(val)):
                        return True
        
        return False
    
    except:
        return True


# ===== RUN BAYESIAN OPTIMIZATION =====
# NOTE: Button state (disabled + children) is managed ONLY by running=[]
# Do NOT add run-bo-btn outputs to the callback outputs, or it will conflict

@callback(
    [Output('experiment-datatable', 'data', allow_duplicate=True),
     Output('bo-result-alert', 'children'),
     Output('bo-result-alert', 'is_open'),
     Output('bo-result-alert', 'color')],
    Input('run-bo-btn', 'n_clicks'),
    [State('experiment-datatable', 'data'),
     State('current-excel-file', 'data'),
     State('advanced-bo-settings-store', 'data')],
    prevent_initial_call=True,
    running=[
        (Output('run-bo-btn', 'disabled'), True, False),
        (Output('run-bo-btn', 'children'), 
         [dbc.Spinner(size="sm", spinner_class_name="me-2"), "Computing..."], 
         [html.I(className="bi bi-lightning-charge me-2"), "Get New Experiment"])
    ]
)
def run_bayesian_optimization(n_clicks, table_data, excel_file, advanced_settings):
    """Run Bayesian optimization and add new suggested experiment"""
    
    if not n_clicks:
        raise PreventUpdate
    
    print("🚀 Starting Bayesian Optimization...")
    
    try:
        # === SAFE AUTO-SAVE before running optimization ===
        file_path = os.path.join(EXCEL_FOLDER, excel_file)
        df_to_save = pd.DataFrame(table_data)
        
        success, save_msg = safe_excel_save(
            file_path,
            lambda path: df_to_save.to_excel(path, index=False, engine='openpyxl')
        )
        if not success:
            return no_update, f"❌ Auto-save failed: {save_msg}", True, "danger"
        print(f"💾 Table auto-saved to {excel_file}")
        
        # Load domain
        domain_data = DomainStorage.load_domain(excel_file)
        if not domain_data:
            return no_update, "❌ Domain configuration not found", True, "danger"
        
        param_names = domain_data.get('metadata', {}).get('parameter_names', [])
        obj_names = domain_data.get('metadata', {}).get('objective_names', [])
        parameters = domain_data.get('parameters', [])
        objectives = domain_data.get('objectives', [])
        
        solvent_config = domain_data.get('metadata', {}).get('solvent_config')
        base_config = domain_data.get('metadata', {}).get('base_config')
        constraints_config = domain_data.get('metadata', {}).get('constraints_config')

        # Create domain
        discretization_config = {}
    
        for param in parameters:
            param_name = param.get('name')
            param_type = param.get('type')
            param_type_info = param.get('type_info', {})
            
            if 'step' in param_type_info:
                step_value = param_type_info.get('step', 0)
                if step_value and step_value > 0:
                    discretization_config[param_name] = float(step_value)
                    print(f"   🎯 '{param_name}' discretized (stored step={step_value})")
                    continue
        
        if discretization_config:
            print(f"🎯 Reconstructed discretization config: {discretization_config}")
        else:
            print(f"ℹ️ No discretization for optimization (continuous parameters)")
        
        # Recreate domain WITH discretization
        domain = create_bofire_domain_from_store(
            parameters, 
            objectives,
            solvent_config=solvent_config,
            base_config=base_config,
            constraints_config=constraints_config,
            discretization_config=discretization_config
        )
        
        # === SAFE READ for optimization data ===
        df, read_msg = safe_excel_read(file_path)
        if df is None:
            return no_update, f"❌ {read_msg}", True, "danger"
        if "backup" in read_msg:
            print(f"⚠️ {read_msg}")
        
        # ===== FILTRER UNIQUEMENT LES COLONNES DU DOMAINE =====
        valid_columns = param_names + obj_names
        missing_cols = [col for col in valid_columns if col not in df.columns]
        if missing_cols:
            return no_update, f"❌ Missing columns: {missing_cols}", True, "danger"
        
        # Prepare experiments - GARDER UNIQUEMENT LES COLONNES DU DOMAINE
        experiments = df[valid_columns].copy()
        print(f"🔍 Filtered experiments to domain columns: {list(experiments.columns)}")
        
        param_definitions = {p['name']: p for p in parameters}
        
        for col in experiments.columns:
            if col in param_names:
                param_def = param_definitions.get(col)
                if param_def:
                    param_type = param_def.get('type')
                    type_info = param_def.get('type_info', {})
                    
                    if param_type == 'cat':
                        allowed_values = type_info.get('values', [])
                        print(f"🏷️ Categorical parameter '{col}' - allowed values: {allowed_values}")
                        
                        invalid_mask = ~experiments[col].isin(allowed_values + ['', None, 'nan', 'None'])
                        if invalid_mask.any():
                            invalid_vals = experiments.loc[invalid_mask, col].unique()
                            error_msg = f"❌ Invalid values in '{col}': {list(invalid_vals)}. Allowed: {allowed_values}"
                            print(error_msg)
                            return no_update, error_msg, True, "danger"
                        
                        if allowed_values:
                            experiments[col] = experiments[col].replace(['nan', 'None', ''], allowed_values[0])
                        
                    elif param_type == 'int':
                        allowed_values = type_info.get('range', [])
                        print(f"🔢 Discrete parameter '{col}' - allowed values: {allowed_values}")
                        experiments[col] = pd.to_numeric(experiments[col], errors='coerce')
                        
                    elif param_type == 'float':
                        experiments[col] = pd.to_numeric(experiments[col], errors='coerce')
                        
            else:
                experiments[col] = pd.to_numeric(experiments[col], errors='coerce')
        
        for obj in obj_names:
            if obj in experiments.columns:
                nan_count = experiments[obj].isna().sum()
                if nan_count > 0:
                    print(f"⚠️ {nan_count} NaN values in {obj}")
        
        complete_mask = pd.Series([True] * len(experiments))
        for obj in obj_names:
            if obj in experiments.columns:
                complete_mask = complete_mask & experiments[obj].notna()
        
        experiments_complete = experiments[complete_mask].copy()
        
        if len(experiments_complete) == 0:
            return no_update, "❌ No complete experiments found. Please fill in all objective values.", True, "danger"
        
        print(f"📈 Complete experiments data ({len(experiments_complete)} rows):\n{experiments_complete}")
        print(f"📈 Data types:\n{experiments_complete.dtypes}")
        
        # ===== PARAMÈTRES AVANCÉS =====
        advanced_settings = advanced_settings or {}
        acq_func_name = advanced_settings.get('acquisition_function', 'qLogNEI (default)')
        n_suggestions = advanced_settings.get('n_candidates', 1)
    
        # Déterminer type et créer fonction d'acquisition
        opt_type = get_optimization_type(domain)
        is_multi_objective = (opt_type == 'MOBO')
        acq_func = create_acquisition_function_from_name(acq_func_name, is_multi_objective)
    
        # ── Outcome constraint (MOBO uniquement) ──────────────────────────────
        oc = advanced_settings.get('outcome_constraint') or {}
        outcome_constraint = None
        if (is_multi_objective
                and oc.get('enabled')
                and oc.get('objective')
                and oc.get('threshold') is not None):
            outcome_constraint = oc
            print(f"🔒 Outcome constraint active : {oc['objective']} {oc['direction']} {oc['threshold']}")
    
        print(f"🔧 Running {opt_type} with {acq_func_name}, {n_suggestions} candidates")

        new_candidates = bayesian_optimization(
            domain,
            experiments_complete,
            n_candidates=n_suggestions,
            acquisition_function=acq_func,
            outcome_constraint=outcome_constraint,
        )
        
        if new_candidates is None or (hasattr(new_candidates, 'empty') and new_candidates.empty):
            print("❌ No candidates generated")
            return no_update, "❌ Optimization failed to generate candidates", True, "danger"
        
        print(f"✅ Generated candidates:\n{new_candidates}")
        
        # ===== DYNAMIC BOILING POINT CONSTRAINT =====
        constraints_config = domain_data.get('metadata', {}).get('constraints_config')
        solvent_config = domain_data.get('metadata', {}).get('solvent_config')
        
        # Find solvent parameter name
        solvent_param_name = None
        if solvent_config:
            for param in parameters:
                if param.get('id') == solvent_config.get('param_id'):
                    solvent_param_name = param.get('name')
                    break
        
        # Apply dynamic constraints based on suggested solvent
        all_adjustments = []
        if constraints_config and constraints_config.get('constraints') and solvent_param_name:
            print(f"🔧 Checking dynamic constraints (solvent param: {solvent_param_name})")
            
            adjusted_rows = []
            for idx, row in new_candidates.iterrows():
                row_dict = row.to_dict()
                adjusted_row, adjustments = validate_and_adjust_suggestion(
                    row_dict, 
                    constraints_config, 
                    solvent_param_name
                )
                adjusted_rows.append(adjusted_row)
                all_adjustments.extend(adjustments)
            
            new_candidates = pd.DataFrame(adjusted_rows)
            
            if all_adjustments:
                print(f"⚠️ Applied {len(all_adjustments)} temperature adjustment(s):")
                for adj in all_adjustments:
                    print(f"   {adj['parameter']}: {adj['original']}°C → {adj['adjusted']}°C "
                          f"(limited by {adj['solvent']} BP={adj['boiling_point']}°C)")
        
        new_rows = []
        for _, candidate in new_candidates.iterrows():
            new_row = {}
            
            for col in df.columns:
                if col == 'Point type':
                    new_row[col] = 'BO'
                elif col in param_names:
                    val = candidate.get(col, '')
                    param_def = param_definitions.get(col)
                    if param_def:
                        if param_def.get('type') == 'float' and isinstance(val, (int, float)):
                            new_row[col] = round(float(val), 2)
                        elif param_def.get('type') == 'cat':
                            new_row[col] = str(val) if val else ''
                        else:
                            new_row[col] = val
                    else:
                        new_row[col] = val
                elif col in obj_names:
                    new_row[col] = ''
                else:
                    new_row[col] = ''
            
            new_rows.append(new_row)
        
        updated_data = table_data + new_rows
        
        if all_adjustments:
            adjustment_text = ", ".join([
                f"{adj['parameter']} → {adj['adjusted']}°C (max for {adj['solvent']})"
                for adj in all_adjustments
            ])
            msg = f"✅ Generated {n_suggestions} new experiment(s)! ⚠️ Adjusted: {adjustment_text}"
        else:
            msg = f"✅ Generated {n_suggestions} new experiment(s) to test!"
        print(msg)
        return updated_data, msg, True, "success"
    
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"💥 Optimization error:\n{error_trace}")
        
        error_str = str(e)
        if "invalid values" in error_str.lower():
            hint = "\n\n💡 Hint: Check that all categorical parameter values in your data match the allowed categories defined in your domain."
            return no_update, f"❌ Optimization error: {error_str}{hint}", True, "danger"
        
        return no_update, f"❌ Optimization error: {error_str}", True, "danger"


# ===== AUTO-SAVE ON TABLE EDIT =====

@callback(
    Output('auto-save-indicator', 'children'),
    Input('experiment-datatable', 'data_timestamp'),
    [State('experiment-datatable', 'data'),
     State('current-excel-file', 'data')],
    prevent_initial_call=True
)
def auto_save_on_edit(timestamp, table_data, excel_file):
    """Auto-save table when edited"""
    
    if not timestamp or not table_data or not excel_file:
        raise PreventUpdate
    
    from datetime import datetime
    
    file_path = os.path.join(EXCEL_FOLDER, excel_file)
    df = pd.DataFrame(table_data)
    
    # === SAFE SAVE with atomic write + backup ===
    success, msg = safe_excel_save(
        file_path,
        lambda path: df.to_excel(path, index=False, engine='openpyxl')
    )
    
    if success:
        return html.Small(f"Auto-saved at {datetime.now().strftime('%H:%M:%S')}", 
                         className="text-muted")
    else:
        return html.Small(f"⚠️ {msg}", className="text-danger")