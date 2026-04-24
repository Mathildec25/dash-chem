"""
Domain creation and Excel generation with sampling
Main callback for creating domains and generating initial experiments
"""

import dash
from dash import callback, Input, Output, State, ALL, no_update
from dash.exceptions import PreventUpdate
import dash_bootstrap_components as dbc
from dash import html
import pandas as pd
import numpy as np
import os
import uuid

from domain_storage import DomainStorage
from utils.BoFire import create_bofire_domain_from_store
from utils.bofire_optimization import sampling, kmeans_sampling
from utils.safe_excel import safe_excel_save, safe_excel_read
from config_path import EXCEL_FOLDER, TRACKING_FILE
from callbacks.advanced_bo_callbacks import _get_available_campaigns


# =============================================================================
# HELPERS — READ-ONLY DISPLAYS + FRESH EDITABLE CONTAINERS
# =============================================================================

def _make_readonly_param_display(parameters):
    """Compact read-only display for parameters when TL is active."""
    type_labels = {'float': 'Continuous', 'int': 'Discrete', 'cat': 'Categorical'}

    rows = []
    for p in parameters:
        t = p.get('type', '')
        ti = p.get('type_info', {})

        if t in ('float', 'int'):
            r = ti.get('range', [])
            value_str = f"{r[0]} – {r[1]}" if len(r) == 2 else str(r)
            step = ti.get('step')
            if step:
                value_str += f"  (step {step})"
        else:
            value_str = ', '.join(str(v) for v in ti.get('values', []))

        rows.append(
            dbc.Row([
                dbc.Col(html.Span(p['name'], className="fw-bold small"), width=4),
                dbc.Col(dbc.Badge(type_labels.get(t, t), color="secondary",
                                  className="me-1"), width=2),
                dbc.Col(html.Span(value_str, className="text-muted small"), width=6),
            ], className="mb-1 align-items-center")
        )

    return html.Div([
        dbc.Alert([html.I(className="bi bi-lock-fill me-2"),
                   "Parameters locked — imported from source campaign"],
                  color="info", className="py-2 mb-2", style={"fontSize": "0.8rem"}),
        *rows
    ])


def _make_readonly_obj_display(objectives):
    """Compact read-only display for objectives when TL is active."""
    rows = []
    for o in objectives:
        direction = "↑ Maximize" if o.get('direction') == 'max' else "↓ Minimize"
        bounds    = f"[{o.get('lower_bound', 0)} – {o.get('upper_bound', 1)}]"
        rows.append(
            dbc.Row([
                dbc.Col(html.Span(o['name'], className="fw-bold small"), width=5),
                dbc.Col(html.Span(direction, className="text-muted small"), width=4),
                dbc.Col(html.Span(bounds,    className="text-muted small"), width=3),
            ], className="mb-1 align-items-center")
        )

    return html.Div([
        dbc.Alert([html.I(className="bi bi-lock-fill me-2"),
                   "Objectives locked — imported from source campaign"],
                  color="info", className="py-2 mb-2", style={"fontSize": "0.8rem"}),
        *rows
    ])


def _make_fresh_editable_containers():
    """
    Return fresh editable parameter + objective containers (new UUIDs).
    Called when TL is disabled to restore the default empty form.
    """
    param_id = str(uuid.uuid4())
    obj_id   = str(uuid.uuid4())

    param_row = html.Div([
        dbc.Row([
            dbc.Col(dbc.Input(id={'type': 'parameter-name', 'index': param_id},
                              placeholder="Parameter name", size="sm",
                              style={"borderRadius": "6px"}), width=3),
            dbc.Col(dcc.Dropdown(id={'type': 'parameter-type', 'index': param_id},
                                 options=[
                                     {"label": "Continuous", "value": "float"},
                                     {"label": "Discrete",   "value": "int"},
                                     {"label": "Categorical","value": "cat"},
                                 ], value="float", clearable=False,
                                 style={"fontSize": "0.875rem"}), width=2),
            dbc.Col(html.Div(id={'type': 'parameter-inputs', 'index': param_id}, children=[
                dbc.Row([
                    dbc.Col(dbc.Input(id={'type': 'parameter-min', 'index': param_id},
                                      placeholder="Min", type="number", step="any", size="sm",
                                      style={"borderRadius": "6px"}), width=4),
                    dbc.Col(dbc.Input(id={'type': 'parameter-max', 'index': param_id},
                                      placeholder="Max", type="number", step="any", size="sm",
                                      style={"borderRadius": "6px"}), width=4),
                    dbc.Col(dbc.Input(id={'type': 'parameter-step', 'index': param_id},
                                      placeholder="Step", type="number", step="any", size="sm",
                                      style={"borderRadius": "6px"}), width=4),
                ])
            ]), width=5),
            dbc.Col(html.Div(id={'type': 'parameter-categories', 'index': param_id},
                             style={"display": "none"}), width=0),
            dbc.Col(dbc.Button(html.I(className="bi bi-trash",
                                      style={"fontSize": "0.875rem"}),
                               id={'type': 'delete-parameter', 'index': param_id},
                               color="danger", outline=True, size="sm",
                               style={"borderRadius": "6px", "padding": "0.25rem 0.5rem"}),
                    width=1),
        ], className="mb-2 align-items-center"),
    ], id={'type': 'parameter-row', 'index': param_id})

    obj_row = html.Div([
        dbc.Row([
            dbc.Col(dbc.Input(id={'type': 'objective-name', 'index': obj_id},
                              placeholder="Objective name", size="sm",
                              style={"borderRadius": "6px"}), width=4),
            dbc.Col(dcc.Dropdown(id={'type': 'objective-direction', 'index': obj_id},
                                 options=[{"label": "Minimize", "value": "min"},
                                          {"label": "Maximize", "value": "max"}],
                                 placeholder="Direction", clearable=False,
                                 style={"fontSize": "0.875rem"}), width=2),
            dbc.Col(dbc.Input(id={'type': 'objective-lower', 'index': obj_id},
                              placeholder="Min", type="number", step="any", size="sm",
                              style={"borderRadius": "6px"}), width=2),
            dbc.Col(dbc.Input(id={'type': 'objective-upper', 'index': obj_id},
                              placeholder="Max", type="number", step="any", size="sm",
                              style={"borderRadius": "6px"}), width=2),
            dbc.Col(dbc.Button(html.I(className="bi bi-trash",
                                      style={"fontSize": "0.875rem"}),
                               id={'type': 'delete-objective', 'index': obj_id},
                               color="danger", outline=True, size="sm",
                               style={"borderRadius": "6px", "padding": "0.25rem 0.5rem"}),
                    width=2),
        ], className="mb-1 align-items-center"),
    ], id={'type': 'objective-row', 'index': obj_id})

    return [param_row], [obj_row]


# We need dcc.Dropdown in the fresh containers — import it here
from dash import dcc


# =============================================================================
# TL — POPULATE SOURCE DROPDOWN ON PAGE LOAD
# =============================================================================

@callback(
    Output('tl-domain-source-dropdown', 'options'),
    Input('url', 'pathname'),
    prevent_initial_call=False
)
def populate_tl_source_options(pathname):
    """Populate TL source dropdown with all existing campaigns."""
    if pathname != '/Opt-param':
        raise PreventUpdate
    return _get_available_campaigns(current_excel_file=None)


# =============================================================================
# TL — MAIN CONFIGURATION CALLBACK
# Handles both switch toggle and source selection in one place.
# =============================================================================

@callback(
    [Output('parameter-container',          'children', allow_duplicate=True),
     Output('objective-container',          'children', allow_duplicate=True),
     Output('tl-domain-store',              'data'),
     Output('param-action-buttons',         'style'),
     Output('add-objective-button',         'style'),
     Output('tl-domain-source-section',     'style'),
     Output('sampling-config-row',          'style'),
     Output('tl-best-point-info',           'style'),
     Output('tl-domain-compatibility-badge','children')],
    [Input('tl-domain-switch',              'value'),
     Input('tl-domain-source-dropdown',     'value')],
    prevent_initial_call=True
)
def on_tl_config_changed(tl_enabled, source_excel):
    """
    Central handler for Transfer Learning UI:
      - Switch OFF or no source selected → restore fresh editable form
      - Switch ON + source selected      → lock form with source data,
                                           hide sampling card, show best-point banner
    """
    tl_enabled   = bool(tl_enabled)
    source_section_style = {"display": "block"} if tl_enabled else {"display": "none"}

    # ── TL off or no source → restore editable form ───────────────────────
    if not tl_enabled or not source_excel:
        param_container, obj_container = _make_fresh_editable_containers()
        return (
            param_container,
            obj_container,
            None,                              # clear store
            {"display": "block"},              # show param buttons
            {"display": "inline-block"},       # show add-obj button
            source_section_style,
            {"display": "block"},              # show sampling card
            {"display": "none"},               # hide banner
            html.Span()                        # clear badge
        )

    # ── TL on + source selected → lock form ───────────────────────────────
    source_data = DomainStorage.load_domain(source_excel)
    if not source_data:
        param_container, obj_container = _make_fresh_editable_containers()
        badge = dbc.Alert([html.I(className="bi bi-x-circle me-2"),
                           "Could not load source campaign"],
                          color="danger", className="py-1 px-2 mb-0",
                          style={"fontSize": "0.8rem"})
        return (
            param_container, obj_container, None,
            {"display": "block"}, {"display": "inline-block"},
            source_section_style,
            {"display": "block"}, {"display": "none"},
            badge
        )

    parameters = source_data.get('parameters', [])
    objectives  = source_data.get('objectives', [])

    # Count available complete experiments in source
    try:
        source_path  = os.path.join(EXCEL_FOLDER, source_excel)
        df_src       = pd.read_excel(source_path, engine='openpyxl')
        obj_names    = [o['name'] for o in objectives]
        n_complete   = df_src.dropna(subset=obj_names).shape[0]

        # Identify best condition for the banner
        first_obj    = objectives[0] if objectives else None
        best_desc    = ""
        if first_obj and n_complete > 0:
            ascending  = (first_obj['direction'] == 'min')
            df_complete = df_src.dropna(subset=obj_names).copy()
            df_complete[first_obj['name']] = pd.to_numeric(
                df_complete[first_obj['name']], errors='coerce')
            best_row   = df_complete.sort_values(
                first_obj['name'], ascending=ascending).iloc[0]
            best_val   = best_row[first_obj['name']]
            best_desc  = f" ({first_obj['name']}={best_val:.3g})"

        badge = dbc.Alert(
            [html.I(className="bi bi-check-circle me-2"),
             f"Compatible ✅  ({n_complete} experiments){best_desc}"],
            color="success", className="py-1 px-2 mb-0",
            style={"fontSize": "0.8rem"}
        )
    except Exception:
        badge = dbc.Alert([html.I(className="bi bi-check-circle me-2"), "Compatible ✅"],
                          color="success", className="py-1 px-2 mb-0",
                          style={"fontSize": "0.8rem"})

    # Exclude 'domain' (BoFire object, not JSON-serializable) before storing
    store_data = {k: v for k, v in source_data.items() if k != 'domain'}

    return (
        [_make_readonly_param_display(parameters)],
        [_make_readonly_obj_display(objectives)],
        store_data,                            # only JSON-serializable keys
        {"display": "none"},                   # hide param buttons
        {"display": "none"},                   # hide add-obj button
        source_section_style,
        {"display": "none"},                   # hide sampling card
        {"display": "block"},                  # show banner
        badge
    )


# =============================================================================
# ENABLE / DISABLE CREATE BUTTON
# =============================================================================

@callback(
    Output('create-domain-btn', 'disabled'),
    [Input({'type': 'parameter-name', 'index': ALL}, 'value'),
     Input({'type': 'objective-name',  'index': ALL}, 'value'),
     Input('project-name-store', 'data'),
     Input('tl-domain-store',    'data')],
    State('tl-domain-switch', 'value'),
    prevent_initial_call=True
)
def enable_button(param_names, obj_names, project_name, tl_store, tl_enabled):
    """Enable button when requirements are met — handles both manual and TL modes."""

    if not project_name or not project_name.strip():
        return True

    # TL mode: button enabled as soon as a compatible source is loaded
    if bool(tl_enabled) and tl_store:
        return False

    # Manual mode: need at least one named parameter and one named objective
    valid_params = [p for p in param_names if p and p.strip()]
    if not valid_params:
        return True

    valid_objs = [o for o in obj_names if o and o.strip()]
    if not valid_objs:
        return True

    return False


# =============================================================================
# MAIN DOMAIN CREATION CALLBACK
# =============================================================================

@callback(
    [Output('current-excel-file', 'data'),
     Output('url', 'pathname', allow_duplicate=True),
     Output('validation-alert', 'children'),
     Output('validation-alert', 'is_open')],
    Input('create-domain-btn', 'n_clicks'),
    [State('project-name-store', 'data'),
     State({'type': 'parameter-name', 'index': ALL}, 'id'),
     State({'type': 'parameter-name', 'index': ALL}, 'value'),
     State({'type': 'parameter-type', 'index': ALL}, 'value'),
     State({'type': 'parameter-min',  'index': ALL}, 'id'),
     State({'type': 'parameter-min',  'index': ALL}, 'value'),
     State({'type': 'parameter-max',  'index': ALL}, 'id'),
     State({'type': 'parameter-max',  'index': ALL}, 'value'),
     State({'type': 'parameter-categories', 'index': ALL}, 'id'),
     State({'type': 'parameter-categories', 'index': ALL}, 'value'),
     State({'type': 'parameter-step', 'index': ALL}, 'id'),
     State({'type': 'parameter-step', 'index': ALL}, 'value'),
     State({'type': 'objective-name',      'index': ALL}, 'id'),
     State({'type': 'objective-name',      'index': ALL}, 'value'),
     State({'type': 'objective-direction', 'index': ALL}, 'value'),
     State({'type': 'objective-lower',     'index': ALL}, 'value'),
     State({'type': 'objective-upper',     'index': ALL}, 'value'),
     State({'type': 'extra-column-name',   'index': ALL}, 'id'),
     State({'type': 'extra-column-name',   'index': ALL}, 'value'),
     State('starting-sampling-DD',  'value'),
     State('nb-sampling-points',    'value'),
     State('solvent-config-store',  'data'),
     State('base-config-store',     'data'),
     State('constraints-store',     'data'),
     # ── Transfer Learning ──────────────────────────────────────────────
     State('tl-domain-switch',          'value'),
     State('tl-domain-source-dropdown', 'value'),
     State('tl-domain-store',           'data')],
    prevent_initial_call=True
)
def create_domain_and_excel(
        n_clicks, project_name,
        param_ids, param_names, param_types,
        min_ids, param_mins, max_ids, param_maxs,
        cat_ids, param_cats,
        step_ids, param_steps,
        obj_ids, obj_names, obj_directions, obj_lowers, obj_uppers,
        extra_ids, extra_names,
        sampling_method, nb_points,
        solvent_config, base_config, constraints_config,
        tl_enabled, tl_source_campaign, tl_store):
    """
    Create domain, generate Excel, and redirect.

    When Transfer Learning is active:
      - Parameters / objectives are read from tl_store (source domain),
        not from the (locked, empty) form inputs.
      - Initial sampling is replaced by a single point: the best-performing
        condition from the source campaign (Summit strategy).
      - TL config is persisted under 'advanced_bo_settings' in the domain
        metadata so load_settings_from_metadata picks it up automatically.
    """

    if not n_clicks:
        raise PreventUpdate

    tl_enabled = bool(tl_enabled)

    try:

        # ── BRANCH: TL mode ──────────────────────────────────────────────
        if tl_enabled and tl_store and tl_source_campaign:
            parameters = tl_store.get('parameters', [])
            objectives  = tl_store.get('objectives', [])
            print(f"TL mode: {len(parameters)} params, {len(objectives)} objs "
                  f"from '{tl_source_campaign}'")

        # ── BRANCH: Manual mode ──────────────────────────────────────────
        else:
            tl_enabled         = False
            tl_source_campaign = None

            cats_dict  = {cid['index']: cval for cid, cval in zip(cat_ids, param_cats)}
            mins_dict  = {mid['index']: mval for mid, mval in zip(min_ids, param_mins)}
            maxs_dict  = {mid['index']: mval for mid, mval in zip(max_ids, param_maxs)}
            steps_dict = {sid['index']: sval for sid, sval in zip(step_ids, param_steps)
                          if sval is not None}

            parameters = []
            for pid, name, ptype in zip(param_ids, param_names, param_types):
                if not name or not name.strip():
                    continue
                idx = pid['index']

                if ptype == 'float':
                    lb, ub = mins_dict.get(idx), maxs_dict.get(idx)
                    if lb is None or ub is None:
                        alert = dbc.Alert(f"❌ Parameter '{name}' missing min or max value",
                                          color="danger")
                        return no_update, no_update, alert, True
                    type_info = {'range': [float(lb), float(ub)]}
                    if idx in steps_dict:
                        type_info['step'] = float(steps_dict[idx])
                    parameters.append({'id': idx, 'name': name.strip(),
                                       'type': 'float', 'type_info': type_info})

                elif ptype == 'int':
                    cats = cats_dict.get(idx)
                    if not cats:
                        alert = dbc.Alert(
                            f"❌ Discrete parameter '{name}' needs values (e.g., 1, 2, 3)",
                            color="danger")
                        return no_update, no_update, alert, True
                    values = []
                    for v in str(cats).split(','):
                        v = v.strip()
                        if v:
                            try:
                                values.append(float(v))
                            except ValueError as e:
                                alert = dbc.Alert(
                                    f"❌ Discrete parameter '{name}' invalid values: {e}",
                                    color="danger")
                                return no_update, no_update, alert, True
                    if not values:
                        alert = dbc.Alert(
                            f"❌ Discrete parameter '{name}' needs at least one value",
                            color="danger")
                        return no_update, no_update, alert, True
                    parameters.append({'id': idx, 'name': name.strip(),
                                       'type': 'int', 'type_info': {'range': values}})

                elif ptype == 'cat':
                    cats = cats_dict.get(idx)
                    if not cats:
                        alert = dbc.Alert(
                            f"❌ Categorical parameter '{name}' needs values (e.g., A, B, C)",
                            color="danger")
                        return no_update, no_update, alert, True
                    values = [v.strip() for v in str(cats).split(',') if v.strip()]
                    if not values:
                        alert = dbc.Alert(
                            f"❌ Categorical parameter '{name}' needs at least one value",
                            color="danger")
                        return no_update, no_update, alert, True
                    parameters.append({'id': idx, 'name': name.strip(),
                                       'type': 'cat', 'type_info': {'values': values}})

            if not parameters:
                alert = dbc.Alert("❌ At least one valid parameter is required", color="danger")
                return no_update, no_update, alert, True

            objectives = []
            for i, (oid, name, direction) in enumerate(
                    zip(obj_ids, obj_names, obj_directions)):
                if not name or not name.strip() or not direction:
                    continue
                objectives.append({
                    'id': oid['index'], 'name': name.strip(), 'direction': direction,
                    'lower_bound': obj_lowers[i] if i < len(obj_lowers) and obj_lowers[i] is not None else 0.0,
                    'upper_bound': obj_uppers[i] if i < len(obj_uppers) and obj_uppers[i] is not None else 1.0,
                })

        print(f"{len(parameters)} parameters, {len(objectives)} objectives")

        # ── Extra columns ─────────────────────────────────────────────────
        extra_columns = []
        if extra_ids and extra_names:
            for eid, name in zip(extra_ids, extra_names):
                if name and name.strip():
                    extra_columns.append({'id': eid['index'], 'name': name.strip()})
        extra_columns.append({'id': str(uuid.uuid4()), 'name': 'Point type'})

        # ── Create BoFire domain ──────────────────────────────────────────
        discretization_config = {}
        for param in parameters:
            if param.get('type') == 'float' and 'step' in param.get('type_info', {}):
                discretization_config[param['name']] = float(param['type_info']['step'])

        try:
            domain = create_bofire_domain_from_store(
                parameters, objectives,
                solvent_config=solvent_config,
                base_config=base_config,
                constraints_config=constraints_config,
                discretization_config=discretization_config
            )
            print("BoFire domain created")
        except Exception as e:
            import traceback
            print(f"💥 Domain creation error:\n{traceback.format_exc()}")
            alert = dbc.Alert(f"❌ Failed to create domain: {str(e)}", color="danger")
            return no_update, no_update, alert, True

        # ── Excel filename ────────────────────────────────────────────────
        excel_name = project_name.strip()
        if not excel_name.endswith('.xlsx'):
            excel_name += '.xlsx'
        file_path = os.path.join(EXCEL_FOLDER, excel_name)

        # ── Column order ──────────────────────────────────────────────────
        all_columns = (
            [{'name': c['name'], 'type': 'extra'} for c in extra_columns]
            + [{'name': p['name'], 'type': 'parameter', 'data': p} for p in parameters]
            + [{'name': o['name'], 'type': 'objective'} for o in objectives]
        )

        # ─────────────────────────────────────────────────────────────────
        #  INITIAL EXPERIMENT GENERATION
        # ─────────────────────────────────────────────────────────────────
        # Use tl_store excel_name as authoritative fallback: the dropdown
        # State value (tl_source_campaign) can occasionally be None if
        # Dash serialises the hidden section before the callback completes.
        effective_source = tl_source_campaign or (
            tl_store.get('excel_name') if tl_store else None
        )

        if tl_enabled and tl_store and effective_source:
            # ── TL: use best condition from source campaign ───────────────
            print(f"TL init: loading best point from '{effective_source}'")
            source_path = os.path.join(EXCEL_FOLDER, effective_source)
            df_source, _ = safe_excel_read(source_path)

            sampled_data = None
            if df_source is not None:
                obj_names_list   = [o['name'] for o in objectives]
                param_names_list = [p['name'] for p in parameters]
                missing = [c for c in param_names_list + obj_names_list
                           if c not in df_source.columns]

                if not missing:
                    df_src_complete = df_source.dropna(subset=obj_names_list).copy()
                    for obj in obj_names_list:
                        df_src_complete[obj] = pd.to_numeric(
                            df_src_complete[obj], errors='coerce')
                    df_src_complete = df_src_complete.dropna(subset=obj_names_list)

                    if len(df_src_complete) > 0:
                        first_obj = objectives[0]
                        ascending = (first_obj['direction'] == 'min')
                        best_row  = df_src_complete.sort_values(
                            first_obj['name'], ascending=ascending).iloc[0]

                        sampled_data = pd.DataFrame([{
                            p: best_row[p] for p in param_names_list
                            if p in best_row.index
                        }])
                        print(f"Best source condition: {sampled_data.to_dict('records')[0]}")
                    else:
                        print("TL: no complete experiments in source — starting empty")
                else:
                    print(f"TL: source missing columns {missing} — starting empty")
            else:
                print(f"TL: could not read source file '{source_path}'")

            # Inherit solvent/base/constraints configs from source if not
            # set by the current form (which is locked in TL mode)
            source_meta      = tl_store.get('metadata', {})
            solvent_config   = solvent_config   or source_meta.get('solvent_config')
            base_config      = base_config      or source_meta.get('base_config')
            constraints_config = constraints_config or source_meta.get('constraints_config')
            tl_source_campaign = effective_source
        else:
            # ── Standard sampling ─────────────────────────────────────────
            sampled_data = None
            if sampling_method and sampling_method != 'none' \
                    and nb_points and int(nb_points) > 0:
                try:
                    if sampling_method == 'kmeans':
                        sampled_data = kmeans_sampling(
                            domain=domain, nb_points=int(nb_points),
                            constraints_config=constraints_config)
                    else:
                        method_map = {'random': 'UNIFORM',
                                      'latin_hypercube': 'LHS', 'sobol': 'SOBOL'}
                        sampled_data = sampling(domain, method_map.get(sampling_method, 'LHS'),
                                                int(nb_points))
                    print(f"{len(sampled_data)} sampling points ({sampling_method})")
                except Exception as e:
                    import traceback
                    print(f"Sampling error: {traceback.format_exc()}")
                    alert = dbc.Alert(f"❌ Sampling failed: {str(e)}", color="danger")
                    return no_update, no_update, alert, True

        # ── Build DataFrame ───────────────────────────────────────────────
        if sampled_data is not None and not sampled_data.empty:
            df_excel = pd.DataFrame(index=range(len(sampled_data)))
            for col_info in all_columns:
                col_name = col_info['name']
                if col_info['type'] == 'extra':
                    df_excel[col_name] = 'Init' if col_name == 'Point type' else ''
                elif col_info['type'] == 'parameter':
                    if col_name in sampled_data.columns:
                        vals = sampled_data[col_name].values
                        if col_info.get('data', {}).get('type') == 'float':
                            df_excel[col_name] = [round(v, 2) if pd.notna(v) else v
                                                  for v in vals]
                        else:
                            df_excel[col_name] = vals
                    else:
                        df_excel[col_name] = ''
                elif col_info['type'] == 'objective':
                    df_excel[col_name] = ''
        else:
            headers  = [c['name'] for c in all_columns]
            df_excel = pd.DataFrame(columns=headers)
            empty_row = {c['name']: ('BO' if c['name'] == 'Point type' else '')
                         for c in all_columns}
            df_excel = pd.concat([df_excel, pd.DataFrame([empty_row])], ignore_index=True)

        # ── Save Excel ────────────────────────────────────────────────────
        os.makedirs(EXCEL_FOLDER, exist_ok=True)

        def write_new_excel(path):
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df_excel.to_excel(writer, index=False, sheet_name='Experiments')
                from openpyxl.styles import Font, PatternFill, Alignment
                ws = writer.sheets['Experiments']
                color_map = {'extra': "6C757D", 'parameter': "007BFF", 'objective': "28A745"}
                for i, cell in enumerate(ws[1]):
                    t = all_columns[i]['type']
                    if all_columns[i]['name'] == 'Point type':
                        color = "FF6B35"
                    else:
                        color = color_map.get(t, "6C757D")
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill      = PatternFill(start_color=color, end_color=color,
                                                 fill_type="solid")

        save_ok, save_msg = safe_excel_save(file_path, write_new_excel, backup=False)
        if not save_ok:
            alert = dbc.Alert(f"❌ Excel save failed: {save_msg}", color="danger")
            return no_update, no_update, alert, True
        print(f"Excel saved: {file_path}")

        # ── Save domain metadata ──────────────────────────────────────────
        # TL config stored under 'advanced_bo_settings' so that
        # load_settings_from_metadata() on the Run page picks it up automatically.
        n_obj       = len(objectives)
        default_acq = 'qLogNEHVI (default)' if n_obj >= 2 else 'qLogNEI (default)'

        advanced_bo_settings = {
            'acquisition_function': default_acq,
            'n_candidates': 1,
            'outcome_constraint': {
                'enabled': False, 'objective': None,
                'direction': '>=', 'threshold': None,
            },
            'transfer_learning': {
                'enabled':         tl_enabled,
                'source_campaign': tl_source_campaign if tl_enabled else None,
            }
        }

        success, message = DomainStorage.save_domain(
            excel_name=excel_name,
            domain=domain,
            parameters=parameters,
            objectives=objectives,
            extra_columns=extra_columns,
            metadata={
                'sampling_method':      sampling_method or 'none',
                'nb_points':            nb_points if nb_points else 0,
                'column_order':         [c['name'] for c in all_columns],
                'parameter_names':      [p['name'] for p in parameters],
                'objective_names':      [o['name'] for o in objectives],
                'extra_column_names':   [c['name'] for c in extra_columns],
                'solvent_config':       solvent_config,
                'base_config':          base_config,
                'constraints_config':   constraints_config,
                'advanced_bo_settings': advanced_bo_settings,
            }
        )

        if not success:
            alert = dbc.Alert(f"❌ Domain save failed: {message}", color="danger")
            return no_update, no_update, alert, True

        tl_note = f" with TL from '{tl_source_campaign}'" if tl_enabled else ""
        print(f"Domain saved{tl_note}")

        # ── Update tracking ───────────────────────────────────────────────
        if os.path.exists(TRACKING_FILE):
            df_track, _ = safe_excel_read(TRACKING_FILE)
            if df_track is None:
                df_track = pd.DataFrame(columns=['filename'])
        else:
            df_track = pd.DataFrame(columns=['filename'])

        if excel_name not in df_track['filename'].values:
            df_track = pd.concat(
                [df_track, pd.DataFrame([{'filename': excel_name}])],
                ignore_index=True
            )
            safe_excel_save(TRACKING_FILE,
                            lambda p: df_track.to_excel(p, index=False, engine='openpyxl'))

        # ── Redirect ──────────────────────────────────────────────────────
        alert = dbc.Alert([
            html.I(className="bi bi-check-circle-fill me-2"),
            f"✅ Domain created{tl_note}! Redirecting…"
        ], color="success")

        return excel_name, '/Opt-run', alert, True

    except Exception as e:
        import traceback
        print(f"💥 Error in domain creation:\n{traceback.format_exc()}")
        alert = dbc.Alert([
            html.H6("❌ Creation Failed", className="alert-heading"),
            html.P(str(e))
        ], color="danger")
        return no_update, no_update, alert, True