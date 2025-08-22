import os
import pickle
import uuid
import pandas as pd
from typing import Optional, Dict, Any, Tuple
import json
from datetime import datetime
import numpy as np
from collections import Counter
from matplotlib import pyplot as plt
import dash
from dash import html, dash_table, dcc
import dash_bootstrap_components as dbc

from utils.BoFire import create_bofire_domain_from_store, sampling, optimization
from config_path import EXCEL_FOLDER, DOMAIN_FOLDER, TRACKING_FILE, DOMAIN_TRACKING_FILE, DOMAIN_TRACKING_FILENAME

class DomainStorage:
    @staticmethod
    def save_domain(excel_name: str, domain, parameters: list, objectives: list, 
                   extra_columns: list = None, metadata: dict = None) -> Tuple[bool, str]:
        """
        Save a BoFire domain and associated metadata
        
        Args:
            excel_name: Name of the Excel file (will add .xlsx if missing)
            domain: BoFire domain object
            parameters: List of parameter dictionaries
            objectives: List of objective dictionaries  
            extra_columns: List of extra column dictionaries
            metadata: Additional metadata dictionary
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        try:
            # Ensure folders exist
            os.makedirs(EXCEL_FOLDER, exist_ok=True)
            os.makedirs(DOMAIN_FOLDER, exist_ok=True)

            # Ensure Excel name has .xlsx extension
            if not excel_name.endswith(".xlsx"):
                excel_name += ".xlsx"

            # Save domain object as pickle
            domain_filename = excel_name.replace('.xlsx', '_domain.pkl')
            domain_path = os.path.join(DOMAIN_FOLDER, domain_filename)
            with open(domain_path, 'wb') as f:
                pickle.dump(domain, f)

            # Prepare domain metadata for tracking
            domain_metadata = {
                'excel_file': excel_name,
                'domain_file': domain_filename,
                'created_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'parameters': parameters or [],
                'objectives': objectives or [],
                'extra_columns': extra_columns or [],
                'metadata': metadata or {}
            }

            # Update tracking file
            DomainStorage._update_tracking_file(excel_name, domain_metadata)

            return True, f"Domain saved successfully for {excel_name}"

        except Exception as e:
            return False, f"Failed to save domain: {str(e)}"
    
    @staticmethod
    def load_domain(excel_name: str) -> Tuple[bool, Any, Optional[Dict]]:
        """
        Load a BoFire domain and its metadata for a given Excel file
        
        Args:
            excel_name: Name of the associated Excel file
            
        Returns:
            tuple: (success: bool, domain_object or error_message, metadata_dict or None)
        """
        try:
            # Ensure .xlsx extension
            if not excel_name.endswith(".xlsx"):
                excel_name += ".xlsx"
                
            # Get domain info from tracking file
            domain_info = DomainStorage._get_domain_info(excel_name)
            if not domain_info:
                return False, f"No domain found for {excel_name}", None
            
            # Load domain object
            domain_path = os.path.join(DOMAIN_FOLDER, domain_info['domain_file'])
            if not os.path.exists(domain_path):
                return False, f"Domain file not found: {domain_info['domain_file']}", None
            
            with open(domain_path, 'rb') as f:
                domain = pickle.load(f)
            
            return True, domain, domain_info
            
        except Exception as e:
            return False, f"Failed to load domain: {str(e)}", None
    
    @staticmethod
    def domain_exists(excel_name: str) -> bool:
        """Check if a domain exists for the given Excel file"""
        if not excel_name.endswith(".xlsx"):
            excel_name += ".xlsx"
        return DomainStorage._get_domain_info(excel_name) is not None
    
    @staticmethod
    def list_domains() -> pd.DataFrame:
        """List all stored domains"""
        if os.path.exists(DOMAIN_TRACKING_FILE):
            try:
                return pd.read_excel(DOMAIN_TRACKING_FILE, engine='openpyxl')
            except:
                return pd.DataFrame()
        return pd.DataFrame()
    
    @staticmethod
    def delete_domain(excel_name: str) -> Tuple[bool, str]:
        """Delete a domain and its metadata"""
        try:
            if not excel_name.endswith(".xlsx"):
                excel_name += ".xlsx"
                
            domain_info = DomainStorage._get_domain_info(excel_name)
            if domain_info:
                # Delete domain file
                domain_path = os.path.join(DOMAIN_FOLDER, domain_info['domain_file'])
                if os.path.exists(domain_path):
                    os.remove(domain_path)
                
                # Remove from tracking file
                if os.path.exists(DOMAIN_TRACKING_FILE):
                    df = pd.read_excel(DOMAIN_TRACKING_FILE, engine='openpyxl')
                    df = df[df['excel_file'] != excel_name]
                    df.to_excel(DOMAIN_TRACKING_FILE, index=False, engine='openpyxl')
            
            return True, f"Domain deleted for {excel_name}"
            
        except Exception as e:
            return False, f"Failed to delete domain: {str(e)}"
    
    @staticmethod
    def _update_tracking_file(excel_name: str, domain_metadata: dict):
        """Update the domain tracking file with new metadata"""
        try:
            # Ensure save folder exists
            os.makedirs(EXCEL_FOLDER, exist_ok=True)

            # Load existing tracking data
            if os.path.exists(DOMAIN_TRACKING_FILE):
                df = pd.read_excel(DOMAIN_TRACKING_FILE, engine='openpyxl')
            else:
                df = pd.DataFrame()

            # Remove existing entry for this excel file if it exists
            if not df.empty and 'excel_file' in df.columns:
                df = df[df['excel_file'] != excel_name]

            # Create new row
            new_row = {
                'excel_file': domain_metadata['excel_file'],
                'domain_file': domain_metadata['domain_file'],
                'created_date': domain_metadata['created_date'],
                'parameters_json': json.dumps(domain_metadata['parameters']),
                'objectives_json': json.dumps(domain_metadata['objectives']),
                'extra_columns_json': json.dumps(domain_metadata['extra_columns']),
                'metadata_json': json.dumps(domain_metadata['metadata'])
            }

            # Add new row
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            # Save tracking file
            df.to_excel(DOMAIN_TRACKING_FILE, index=False, engine='openpyxl')

        except Exception as e:
            print(f"❌ Failed to update tracking file: {str(e)}")
    
    @staticmethod
    def _get_domain_info(excel_name: str) -> Optional[Dict[str, Any]]:
        """Get domain information for a given Excel file"""
        try:
            if not os.path.exists(DOMAIN_TRACKING_FILE):
                return None
            
            df = pd.read_excel(DOMAIN_TRACKING_FILE, engine='openpyxl')
            if df.empty or 'excel_file' not in df.columns:
                return None
            
            matches = df[df['excel_file'] == excel_name]
            if matches.empty:
                return None
            
            row = matches.iloc[0]
            return {
                'excel_file': row['excel_file'],
                'domain_file': row['domain_file'],
                'created_date': row['created_date'],
                'parameters': json.loads(row['parameters_json']) if pd.notna(row.get('parameters_json')) else [],
                'objectives': json.loads(row['objectives_json']) if pd.notna(row.get('objectives_json')) else [],
                'extra_columns': json.loads(row['extra_columns_json']) if pd.notna(row.get('extra_columns_json')) else [],
                'metadata': json.loads(row['metadata_json']) if pd.notna(row.get('metadata_json')) else {},
                'parameter_names': json.loads(row['metadata_json'])['parameter_names'] if pd.notna(row.get('metadata_json')) else [],
                'objective_names': json.loads(row['metadata_json'])['objective_names'] if pd.notna(row.get('metadata_json')) else [],
                'extra_column_names': json.loads(row['metadata_json'])['extra_column_names'] if pd.notna(row.get('metadata_json')) else []
            }
        except Exception as e:
            print(f"Error getting domain info: {e}")
            return None


def create_domain_and_excel_with_storage(n_clicks, parameters, objectives, extra_columns, 
                                        excel_name, sampling_method=None, nb_points=None):
    """
    Create BoFire domain and Excel file with proper column organization and sampling
    """
    if not n_clicks:
        return dash.no_update, dash.no_update
        
    try:
        # 1️⃣ Validate inputs
        if not excel_name or not excel_name.strip():
            return create_error_message("Please provide a valid Excel file name."), None
            
        if not parameters:
            return create_error_message("Please define at least one parameter."), None
            
        if not objectives:
            return create_error_message("Please define at least one objective."), None

        # 2️⃣ Create the BoFire domain
        domain = create_bofire_domain_from_store(parameters, objectives)

        # 3️⃣ Prepare Excel filename
        excel_name = excel_name.strip()
        if not excel_name.endswith(".xlsx"):
            excel_name += ".xlsx"
        file_path = os.path.join(EXCEL_FOLDER, excel_name)

        # 4️⃣ AUTOMATICALLY ADD "Point type" COLUMN
        # Ensure extra_columns is a list
        if extra_columns is None:
            extra_columns = []
        
        # Check if "Point type" already exists
        point_type_exists = any(col.get("name") == "Point type" for col in extra_columns)
        
        # Add "Point type" column if it doesn't exist
        if not point_type_exists:
            point_type_column = {
                "id": str(uuid.uuid4()),  # Generate unique ID
                "name": "Point type"
            }
            extra_columns.append(point_type_column)
            print("✅ Automatically added 'Point type' column")

        # 5️⃣ Build column structure in correct order
        # Extra columns → Parameters → Objectives
        column_info = []
        
        # Add extra columns first (including our automatic "Point type")
        if extra_columns:
            for col in extra_columns:
                if col.get("name"):
                    column_info.append({
                        'name': col.get("name"),
                        'type': 'extra',
                        'data': col
                    })
        
        # Add parameter columns
        for param in parameters:
            if param.get("name"):
                column_info.append({
                    'name': param.get("name"),
                    'type': 'parameter',
                    'data': param
                })
        
        # Add objective columns
        for obj in objectives:
            if obj.get("name"):
                column_info.append({
                    'name': obj.get("name"),
                    'type': 'objective', 
                    'data': obj
                })

        if not column_info:
            return create_error_message("No valid columns defined."), None

        # 6️⃣ Handle sampling if requested
        sampled_data = None
        sampling_message = ""
        
        if sampling_method and sampling_method.lower() != "none" and nb_points and int(nb_points) > 0:
            try:
                # Convert dropdown value to correct sampling method
                sampling_mapping = {
                    "random": "UNIFORM",
                    "latin_hypercube": "LHS", 
                    "sobol": "SOBOL"
                }
                
                method_key = sampling_mapping.get(sampling_method.lower())
                if not method_key:
                    raise ValueError(f"Unknown sampling method: {sampling_method}")

                # Generate sampling points
                sampled_data = sampling(domain, method_key, int(nb_points))
                sampling_message = f" with {int(nb_points)} {sampling_method} sampling points"

            except Exception as e:
                return create_error_message(f"Sampling failed: {str(e)}"), None

        # 7️⃣ Create DataFrame with proper structure
        if sampled_data is not None and not sampled_data.empty:
            # Use sampled data as base
            num_rows = len(sampled_data)
            df_excel = pd.DataFrame(index=range(num_rows))
            
            # Fill columns in order
            for col_info in column_info:
                col_name = col_info['name']
                col_type = col_info['type']
                
                if col_type == 'extra':
                    # SPECIAL HANDLING FOR "Point type" COLUMN
                    if col_name == "Point type":
                        # All sampled points are "Init" (initial sampling)
                        df_excel[col_name] = "Init"
                        print(f"✅ Set {num_rows} sampling points as 'Init'")
                    else:
                        # Other extra columns are empty
                        df_excel[col_name] = ""
                elif col_type == 'parameter':
                    if col_name in sampled_data.columns:
                        values = sampled_data[col_name].values
                        
                        # Check parameter type
                        param_def = col_info["data"]
                        ptype = param_def.get("type", "").lower()
                        
                        if ptype == "float":
                            df_excel[col_name] = [round(v, 2) if pd.notna(v) else v for v in values]
                        else:
                            df_excel[col_name] = values
                    else:
                        df_excel[col_name] = ""
                elif col_type == 'objective':
                    # Objective columns are empty (to be filled by user)
                    df_excel[col_name] = ""
        else:
            # No sampling - create empty DataFrame with one row
            headers = [col_info['name'] for col_info in column_info]
            df_excel = pd.DataFrame(columns=headers)
            # Add one empty row for user to start
            empty_row = {}
            for col in headers:
                if col == "Point type":
                    # First manual row gets "BO" (since it's like adding a row)
                    empty_row[col] = "BO"
                else:
                    empty_row[col] = ""
            
            df_excel = pd.concat([df_excel, pd.DataFrame([empty_row])], ignore_index=True)
            print("✅ Created initial empty row with 'BO' point type")

        # 8️⃣ Save Excel file with formatting
        save_formatted_excel_with_point_type_highlighting(df_excel, file_path, column_info)

        # 9️⃣ Save domain with proper metadata (include updated extra_columns)
        success, domain_message = DomainStorage.save_domain(
            excel_name=excel_name,
            domain=domain,
            parameters=parameters,
            objectives=objectives,
            extra_columns=extra_columns,  # This now includes "Point type"
            metadata={
                'sampling_method': sampling_method or "none",
                'nb_points': nb_points if nb_points else 0,
                'column_order': [col_info['name'] for col_info in column_info],
                'parameter_names': [col_info['name'] for col_info in column_info if col_info['type'] == 'parameter'],
                'objective_names': [col_info['name'] for col_info in column_info if col_info['type'] == 'objective'],
                'extra_column_names': [col_info['name'] for col_info in column_info if col_info['type'] == 'extra']
            }
        )
        
        if not success:
            return create_error_message(f"Domain storage failed: {domain_message}"), None

        # 🔟 Update Excel tracking file
        update_excel_tracking(excel_name)

        # 1️⃣1️⃣ Return success (no message needed since user gets redirected)
        return None, excel_name

    except Exception as e:
        return create_error_message(f"Creation failed: {str(e)}"), None


# Optional: Enhanced save_formatted_excel function to highlight Point type column
def save_formatted_excel_with_point_type_highlighting(df_excel: pd.DataFrame, file_path: str, column_info: list):
    """Save Excel file with special formatting for Point type column"""
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_excel.to_excel(writer, index=False, sheet_name='Experiments')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Experiments']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        cell_length = len(str(cell.value)) if cell.value is not None else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add header formatting with color coding
            from openpyxl.styles import Font, PatternFill, Alignment
            
            for i, cell in enumerate(worksheet[1], 1):  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color code by column type
                col_info = next((info for info in column_info if info['name'] == cell.value), None)
                if col_info:
                    if col_info['type'] == 'extra':
                        if cell.value == "Point type":
                            # Special color for Point type column
                            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")  # Orange
                        else:
                            cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")  # Gray
                    elif col_info['type'] == 'parameter':
                        cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")  # Blue
                    elif col_info['type'] == 'objective':
                        cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")  # Green
                else:
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Default gray

            # Optional: Add data validation for Point type column
            if "Point type" in df_excel.columns:
                point_type_col_idx = df_excel.columns.get_loc("Point type") + 1  # Excel is 1-indexed
                point_type_col_letter = worksheet.cell(row=1, column=point_type_col_idx).column_letter
                
                # Add dropdown validation for Point type column
                from openpyxl.worksheet.datavalidation import DataValidation
                dv = DataValidation(type="list", formula1='"Init,BO"', allow_blank=True)
                dv.error = 'Please select Init or BO'
                dv.errorTitle = 'Invalid Point Type'
                worksheet.add_data_validation(dv)
                dv.add(f"{point_type_col_letter}2:{point_type_col_letter}{len(df_excel) + 1}")
                    
    except Exception as e:
        raise Exception(f"Failed to save Excel file: {str(e)}")



def update_excel_tracking(excel_name: str):
    """Update the Excel tracking file"""
    try:
        if os.path.exists(TRACKING_FILE):
            df_tracking = pd.read_excel(TRACKING_FILE, engine='openpyxl')
        else:
            df_tracking = pd.DataFrame(columns=["filename"])
        
        if excel_name not in df_tracking["filename"].values:
            new_row = pd.DataFrame([{"filename": excel_name}])
            df_tracking = pd.concat([df_tracking, new_row], ignore_index=True)
            df_tracking.to_excel(TRACKING_FILE, index=False, engine='openpyxl')
    except Exception as e:
        print(f"Warning: Failed to update Excel tracking: {e}")


def create_error_message(message: str):
    """Create a standardized error message"""
    return html.Div([
        html.H5("❌ Error", className="text-danger"),
        html.P(message, className="text-danger"),
    ])


def prepare_experiments_from_excel_data(experiments_data, metadata) -> Optional[pd.DataFrame]:
    """Convert Excel data to BoFire experiments format"""
    if not experiments_data:
        return None
    
    try:
        # Convert to DataFrame if needed
        if isinstance(experiments_data, dict):
            df = pd.DataFrame(experiments_data)
        elif isinstance(experiments_data, list):
            df = pd.DataFrame(experiments_data)
        elif isinstance(experiments_data, pd.DataFrame):
            df = experiments_data.copy()
        else:
            raise ValueError(f"Unsupported experiments_data type: {type(experiments_data)}")
        
        # Get parameter and objective column names from metadata
        param_names = metadata.get('parameter_names', [])
        obj_names = metadata.get('objective_names', [])
        
        # Combine relevant columns
        relevant_columns = param_names + obj_names
        
        # Filter DataFrame to only include relevant columns that exist
        existing_columns = [col for col in relevant_columns if col in df.columns]
        
        if not existing_columns:
            raise ValueError("No matching columns found between data and domain definition")
        
        # Filter to relevant columns
        experiments_df = df[existing_columns].copy()
        
        # Remove rows where any objective column is NaN or empty
        if obj_names:
            obj_columns_in_df = [col for col in obj_names if col in experiments_df.columns]
            if obj_columns_in_df:
                # Remove rows with NaN, empty string, or None in objective columns
                for obj_col in obj_columns_in_df:
                    experiments_df = experiments_df[
                        (experiments_df[obj_col].notna()) & 
                        (experiments_df[obj_col] != "") &
                        (experiments_df[obj_col] != "None")
                    ]
        
        # Check if we have any complete experiments
        if experiments_df.empty:
            return None
        
        return experiments_df
        
    except Exception as e:
        raise ValueError(f"Failed to prepare experiments data: {str(e)}")


def load_experiments_from_excel_file(excel_name: str, sheet_name: str = None) -> pd.DataFrame:
    """Load experiment data directly from Excel file"""
    try:
        if not excel_name.endswith(".xlsx"):
            excel_name += ".xlsx"
            
        file_path = os.path.join(EXCEL_FOLDER, excel_name)
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {excel_name}")
        
        # Load Excel data
        if excel_name.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            else:
                df = pd.read_excel(file_path, engine='openpyxl')
        
        return df
        
    except Exception as e:
        raise ValueError(f"Failed to load experiments from Excel: {str(e)}")


def check_domain_availability() -> Dict[str, bool]:
    """Check which Excel files have associated domains"""
    domains_df = DomainStorage.list_domains()
    if domains_df.empty:
        return {}
    
    return {row['excel_file']: True for _, row in domains_df.iterrows()}